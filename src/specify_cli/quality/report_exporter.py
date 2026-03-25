"""
Multi-Format Report Export System (Exp 103, 104, 105, 106)

Unified interface for generating quality reports in multiple formats simultaneously.
Provides:
- Console output with result card formatting
- JSON export for CI/CD integration and programmatic access
- HTML reports with interactive charts
- Markdown reports for documentation and git diffs
- CSV export for data analysis and spreadsheet import (Exp 104, 105)
- Excel export with conditional formatting and multiple sheets (Exp 106)

Features:
- Single API call to generate all formats
- Consistent data across all report types
- Result card JSON export for programmatic access
- Batch export with output path management
- Format filtering for selective generation
- Category breakdown in CSV for trend analysis (Exp 104)
- Run metadata for BI integration and trend analysis (Exp 105)
- Excel export with professional styling and conditional formatting (Exp 106)

Usage:
    from specify_cli.quality.report_exporter import export_quality_reports

    # Export all formats
    reports = export_quality_reports(result, output_dir="./reports")

    # Export specific formats only
    reports = export_quality_reports(
        result,
        formats=["json", "html", "csv"],
        output_dir="./reports"
    )

    # Export CSV for analysis (Exp 105: enhanced with run_id, timestamp, metadata)
    reports = export_quality_reports(
        result,
        formats=["csv"],
        output_dir="./reports"
    )
    # CSV includes:
    # - Run metadata (run_id, timestamp, artifact, criteria)
    # - Category scores with run_id for grouping
    # - Gate status and benchmark data
    # - Failed rules details
    # Can be imported into Power BI, Tableau, Excel for visualization and trend analysis
"""

from typing import Dict, Any, List, Optional, Literal, Set
from dataclasses import dataclass, field
from pathlib import Path
from datetime import datetime
import json
import csv
from io import StringIO

# Import existing report generators
from specify_cli.quality.result_card import (
    ResultCardData,
    ResultCardFormatter,
    create_result_card_data,
    format_result_card,
)
from specify_cli.quality.json_report import JSONReportGenerator, generate_json_report
from specify_cli.quality.html_report import HTMLReportGenerator, generate_html_report
# MarkdownReportGenerator and generate_markdown_report are defined at the bottom of this file


# Supported export formats
ReportFormat = Literal["console", "json", "html", "markdown", "csv", "excel"]
ALL_FORMATS: Set[ReportFormat] = {"console", "json", "html", "markdown", "csv", "excel"}


@dataclass
class ExportConfig:
    """Configuration for report export"""
    formats: Set[ReportFormat] = field(default_factory=lambda: ALL_FORMATS.copy())
    output_dir: Optional[str] = None
    filename_prefix: str = "quality_report"
    include_timeline: bool = True
    include_details: bool = True
    compact_console: bool = False
    console_theme: str = "default"
    json_pretty: bool = True
    json_validate: bool = True

    def __post_init__(self):
        """Normalize formats set"""
        # Convert list to set if needed
        if isinstance(self.formats, list):
            self.formats = set(self.formats)
        # Filter to valid formats only
        self.formats = self.formats & ALL_FORMATS


@dataclass
class ExportedReport:
    """Single exported report"""
    format: ReportFormat
    content: str
    path: Optional[str] = None
    size_bytes: int = 0
    generated_at: str = field(default_factory=lambda: datetime.now().isoformat())


@dataclass
class ExportResult:
    """Result of multi-format export operation"""
    reports: Dict[ReportFormat, ExportedReport] = field(default_factory=dict)
    output_dir: Optional[str] = None
    total_size_bytes: int = 0
    generated_at: str = field(default_factory=lambda: datetime.now().isoformat())

    def get_report(self, format: ReportFormat) -> Optional[ExportedReport]:
        """Get specific report format"""
        return self.reports.get(format)

    def get_console_output(self) -> str:
        """Get console-formatted output"""
        console_report = self.reports.get("console")
        return console_report.content if console_report else ""

    def get_json_data(self) -> Dict[str, Any]:
        """Get parsed JSON data from JSON report"""
        json_report = self.reports.get("json")
        if json_report and json_report.content:
            try:
                return json.loads(json_report.content)
            except json.JSONDecodeError:
                return {}
        return {}

    def save_all(self, output_dir: Optional[str] = None) -> None:
        """Save all reports to disk (if not already saved)"""
        target_dir = Path(output_dir or self.output_dir)
        if not target_dir:
            raise ValueError("No output directory specified")

        target_dir.mkdir(parents=True, exist_ok=True)

        for format, report in self.reports.items():
            if format == "console":
                continue  # Console output is not saved

            report_path = target_dir / f"{report.path or f'report.{format}'}"
            report_path.write_text(report.content, encoding="utf-8")
            report.path = str(report_path)


class ReportExporter:
    """
    Unified report exporter for multiple output formats.

    Features:
    - Single entry point for all report generation
    - Consistent data across formats
    - Efficient batch processing
    - Flexible output management
    """

    def __init__(self, config: Optional[ExportConfig] = None):
        """Initialize report exporter

        Args:
            config: Export configuration (uses defaults if not provided)
        """
        self.config = config or ExportConfig()
        self._json_generator = JSONReportGenerator()
        self._html_generator = HTMLReportGenerator()
        self._md_generator = MarkdownReportGenerator()

    def export(
        self,
        result: Dict[str, Any],
        previous_score: Optional[float] = None,
    ) -> ExportResult:
        """Export quality report in configured formats

        Args:
            result: Quality loop result dict
            previous_score: Previous run score for trend calculation

        Returns:
            ExportResult with all generated reports
        """
        export_result = ExportResult(output_dir=self.config.output_dir)

        # Create output directory if specified
        output_path = None
        if self.config.output_dir:
            output_dir = Path(self.config.output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = str(output_dir)

        # Generate console output
        if "console" in self.config.formats:
            console_content = self._generate_console(result, previous_score)
            export_result.reports["console"] = ExportedReport(
                format="console",
                content=console_content,
                size_bytes=len(console_content.encode("utf-8")),
            )

        # Generate JSON report
        if "json" in self.config.formats:
            json_content = self._generate_json(result)
            json_size = len(json_content.encode("utf-8"))

            # Save to file if output directory specified
            json_path = None
            if output_path:
                json_path = Path(output_path) / f"{self.config.filename_prefix}.json"
                json_path.write_text(json_content, encoding="utf-8")
                json_path = str(json_path)

            export_result.reports["json"] = ExportedReport(
                format="json",
                content=json_content,
                path=json_path,
                size_bytes=json_size,
            )

        # Generate HTML report
        if "html" in self.config.formats:
            html_content = self._generate_html(result)
            html_size = len(html_content.encode("utf-8"))

            # Save to file if output directory specified
            html_path = None
            if output_path:
                html_path = Path(output_path) / f"{self.config.filename_prefix}.html"
                html_path.write_text(html_content, encoding="utf-8")
                html_path = str(html_path)

            export_result.reports["html"] = ExportedReport(
                format="html",
                content=html_content,
                path=html_path,
                size_bytes=html_size,
            )

        # Generate Markdown report
        if "markdown" in self.config.formats:
            md_content = self._generate_markdown(result)
            md_size = len(md_content.encode("utf-8"))

            # Save to file if output directory specified
            md_path = None
            if output_path:
                md_path = Path(output_path) / f"{self.config.filename_prefix}.md"
                md_path.write_text(md_content, encoding="utf-8")
                md_path = str(md_path)

            export_result.reports["markdown"] = ExportedReport(
                format="markdown",
                content=md_content,
                path=md_path,
                size_bytes=md_size,
            )

        # Generate CSV report (Exp 104)
        if "csv" in self.config.formats:
            csv_content = self._generate_csv(result)
            csv_size = len(csv_content.encode("utf-8"))

            # Save to file if output directory specified
            csv_path = None
            if output_path:
                csv_path = Path(output_path) / f"{self.config.filename_prefix}.csv"
                csv_path.write_text(csv_content, encoding="utf-8")
                csv_path = str(csv_path)

            export_result.reports["csv"] = ExportedReport(
                format="csv",
                content=csv_content,
                path=csv_path,
                size_bytes=csv_size,
            )

        # Generate Excel report (Exp 106)
        if "excel" in self.config.formats:
            excel_bytes = self._generate_excel_safe(result)

            if excel_bytes is not None:
                excel_size = len(excel_bytes)
                excel_content = f"<Binary Excel data, {excel_size} bytes>"

                # Save to file if output directory specified
                excel_path = None
                if output_path:
                    excel_path = Path(output_path) / f"{self.config.filename_prefix}.xlsx"
                    excel_path.write_bytes(excel_bytes)
                    excel_path = str(excel_path)

                export_result.reports["excel"] = ExportedReport(
                    format="excel",
                    content=excel_content,
                    path=excel_path,
                    size_bytes=excel_size,
                )

        # Calculate total size
        export_result.total_size_bytes = sum(
            r.size_bytes for r in export_result.reports.values()
        )

        return export_result

    def _generate_console(
        self,
        result: Dict[str, Any],
        previous_score: Optional[float] = None,
    ) -> str:
        """Generate console-formatted result card

        Args:
            result: Quality loop result dict
            previous_score: Previous run score for trend

        Returns:
            Console-formatted string
        """
        return format_result_card(
            result,
            previous_score=previous_score,
            compact=self.config.compact_console,
            theme=self.config.console_theme,
        )

    def _generate_json(self, result: Dict[str, Any]) -> str:
        """Generate JSON report with enhanced result card data

        Args:
            result: Quality loop result dict

        Returns:
            JSON string with quality data and result card info
        """
        # Generate standard JSON report
        json_content = generate_json_report(
            result,
            pretty=self.config.json_pretty,
            validate=self.config.json_validate,
        )

        # Parse and add result card data
        try:
            json_data = json.loads(json_content)
        except json.JSONDecodeError:
            # If parsing fails, return as-is
            return json_content

        # Add result card section for programmatic access
        card_data = create_result_card_data(result)
        json_data["result_card"] = {
            "status": card_data.status.value,
            "score": card_data.score,
            "passed": card_data.passed,
            "iteration": f"{card_data.iteration}/{card_data.max_iterations}",
            "phase": card_data.phase,
            "rules": {
                "total": card_data.total_rules,
                "passed": card_data.passed_rules,
                "failed": card_data.failed_rules,
                "warnings": card_data.warnings,
            },
            "duration_seconds": card_data.duration_seconds,
            "priority_profile": card_data.priority_profile,
            "trend_change": card_data.trend_change,
            "gate_status": card_data.gate_status,
            "categories": [
                {
                    "category": cat.category,
                    "failed": cat.failed_count,
                    "warnings": cat.warning_count,
                    "total": cat.total_count,
                    "priority": cat.priority,
                    "sample_rules": cat.sample_rules,
                }
                for cat in card_data.category_summaries
            ],
            "action_items": [
                {
                    "priority": item.priority,
                    "title": item.title,
                    "command": item.command,
                    "description": item.description,
                }
                for item in card_data.action_items
            ],
        }

        return json.dumps(json_data, indent=2 if self.config.json_pretty else None, ensure_ascii=False)

    def _generate_html(self, result: Dict[str, Any]) -> str:
        """Generate HTML report with charts

        Args:
            result: Quality loop result dict

        Returns:
            HTML string
        """
        return generate_html_report(
            result,
            output_path=None,  # Handle path separately
            include_timeline=self.config.include_timeline,
            include_details=self.config.include_details,
        )

    def _generate_markdown(self, result: Dict[str, Any]) -> str:
        """Generate Markdown report

        Args:
            result: Quality loop result dict

        Returns:
            Markdown string
        """
        return generate_markdown_report(
            result,
            output_path=None,  # Handle path separately
            include_timeline=self.config.include_timeline,
            include_details=self.config.include_details,
        )

    def _generate_csv(self, result: Dict[str, Any]) -> str:
        """Generate CSV report with category breakdown for data analysis

        Enhanced for Trend Analysis and BI Integration (Exp 105):
        - Schema version for backwards compatibility
        - Run metadata (timestamp, run_id, artifact, criteria)
        - Gate status and benchmark data
        - BI-friendly field naming

        Args:
            result: Quality loop result dict

        Returns:
            CSV string with tabular quality data
        """
        from uuid import uuid4

        output = StringIO()

        # Extract evaluation data
        evaluation = result.get("state", {}).get("evaluation", {})
        categories = evaluation.get("categories", {})

        # Generate run metadata (Exp 105)
        run_id = result.get("run_id", str(uuid4())[:8])
        timestamp = result.get("timestamp", datetime.now().isoformat())
        artifact = result.get("artifact", "")
        criteria = result.get("criteria", "")
        iterations = result.get("iterations", 0)
        max_iterations = result.get("max_iterations", 0)
        priority_profile = result.get("priority_profile", "default")

        # Gate status (Exp 105)
        gate_result = result.get("gate_result", {})
        gate_status = gate_result.get("passed", None)
        gate_policy = gate_result.get("policy_name", "")

        # Benchmark data (Exp 105)
        benchmark = result.get("benchmark", {})
        benchmark_percentile = benchmark.get("percentile", None)
        benchmark_comparison = benchmark.get("comparison", None)

        # Build multi-section CSV (Exp 105)

        # Section 1: Metadata Header (for BI tools)
        output.write("# Spec Kit Quality Report CSV\n")
        output.write(f"# Schema Version: 1.1\n")
        output.write(f"# Generated: {timestamp}\n")
        output.write(f"# Run ID: {run_id}\n")
        output.write(f"# Artifact: {artifact}\n")
        output.write(f"# Criteria: {criteria}\n")
        output.write("#\n")

        # Section 2: Run Metadata
        output.write("\n# Run Metadata\n")
        fieldnames_meta = [
            "run_id",
            "timestamp",
            "artifact",
            "criteria",
            "iteration",
            "max_iterations",
            "priority_profile",
            "gate_status",
            "gate_policy",
            "benchmark_percentile",
            "benchmark_comparison",
        ]
        writer_meta = csv.DictWriter(output, fieldnames=fieldnames_meta)
        writer_meta.writeheader()

        # Convert gate status to string
        gate_status_str = "PASSED" if gate_status is True else "FAILED" if gate_status is False else "NOT_EVALUATED"

        writer_meta.writerow({
            "run_id": run_id,
            "timestamp": timestamp,
            "artifact": artifact,
            "criteria": criteria,
            "iteration": iterations,
            "max_iterations": max_iterations,
            "priority_profile": priority_profile,
            "gate_status": gate_status_str,
            "gate_policy": gate_policy,
            "benchmark_percentile": f"{benchmark_percentile:.1f}" if benchmark_percentile is not None else "",
            "benchmark_comparison": benchmark_comparison if benchmark_comparison else "",
        })

        # Section 3: Category Scores
        output.write("\n# Category Scores\n")
        fieldnames = [
            "run_id",
            "category",
            "score",
            "passed_rules",
            "total_rules",
            "failed_rules",
            "warnings",
            "priority",
            "weight",
            "weighted_score",
        ]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        # Write summary row first
        overall_score = result.get("score", 0.0)
        passed = result.get("passed", False)
        writer.writerow({
            "run_id": run_id,
            "category": "OVERALL",
            "score": f"{overall_score:.3f}",
            "passed_rules": "N/A",
            "total_rules": "N/A",
            "failed_rules": "N/A",
            "warnings": "N/A",
            "priority": "N/A",
            "weight": "1.0",
            "weighted_score": f"{overall_score:.3f}",
        })

        # Write category rows
        for category_name, category_data in categories.items():
            score = category_data.get("score", 0.0)
            passed_rules = category_data.get("passed", 0)
            total_rules = category_data.get("total", 0)
            failed_rules = category_data.get("failed", 0)
            warnings = category_data.get("warnings", 0)
            priority = category_data.get("priority", "medium")
            weight = category_data.get("weight", 1.0)

            writer.writerow({
                "run_id": run_id,
                "category": category_name,
                "score": f"{score:.3f}",
                "passed_rules": passed_rules,
                "total_rules": total_rules,
                "failed_rules": failed_rules,
                "warnings": warnings,
                "priority": priority,
                "weight": f"{weight:.2f}",
                "weighted_score": f"{score * weight:.3f}",
            })

        # Section 4: Failed Rules Details (Exp 105 enhanced)
        csv_content = output.getvalue()

        # Append failed rules details as fourth section
        failed_rules = evaluation.get("failed_rules", [])
        if failed_rules:
            csv_content += "\n# Failed Rules Details\n"
            fieldnames_rules = [
                "run_id",
                "rule_id",
                "category",
                "severity",
                "message",
                "suggestion",
            ]
            output_rules = StringIO()
            writer_rules = csv.DictWriter(output_rules, fieldnames=fieldnames_rules)
            writer_rules.writeheader()

            for rule in failed_rules:
                writer_rules.writerow({
                    "run_id": run_id,
                    "rule_id": rule.get("id", rule.get("rule_id", "unknown")),
                    "category": rule.get("category", "general"),
                    "severity": rule.get("severity", "fail"),
                    "message": rule.get("message", ""),
                    "suggestion": rule.get("suggestion", ""),
                })

            csv_content += output_rules.getvalue()

        return csv_content

    def _generate_excel(self, result: Dict[str, Any]) -> bytes:
        """Generate Excel report with conditional formatting and multiple sheets

        Features (Exp 106):
        - Multiple sheets: Metadata, Summary, Category Scores, Failed Rules
        - Conditional formatting for visual quality assessment
        - Professional styling with headers and borders
        - Color-coded severity indicators
        - Freeze panes for better navigation
        - Auto-filter for data analysis

        Args:
            result: Quality loop result dict

        Returns:
            Excel file as bytes (for saving to file)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
        except ImportError:
            # If openpyxl is not installed, raise helpful error
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        from uuid import uuid4

        # Create workbook with multiple sheets
        wb = Workbook()

        # Extract evaluation data
        evaluation = result.get("state", {}).get("evaluation", {})
        categories = evaluation.get("categories", {})

        # Generate run metadata
        run_id = result.get("run_id", str(uuid4())[:8])
        timestamp = result.get("timestamp", datetime.now().isoformat())
        artifact = result.get("artifact", "")
        criteria = result.get("criteria", "")
        iterations = result.get("iterations", 0)
        max_iterations = result.get("max_iterations", 0)
        priority_profile = result.get("priority_profile", "default")

        # Gate status
        gate_result = result.get("gate_result", {})
        gate_status = gate_result.get("passed", None)
        gate_policy = gate_result.get("policy_name", "")

        # Benchmark data
        benchmark = result.get("benchmark", {})
        benchmark_percentile = benchmark.get("percentile", None)
        benchmark_comparison = benchmark.get("comparison", None)

        # Overall score
        overall_score = result.get("score", 0.0)
        passed = result.get("passed", False)

        # Define styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title styles
        title_font = Font(bold=True, size=14, color="2F5597")
        subtitle_font = Font(bold=True, size=12)

        # ============================================================================
        # SHEET 1: Metadata
        # ============================================================================
        ws_meta = wb.active
        ws_meta.title = "Metadata"

        # Title
        ws_meta['A1'] = "Spec Kit Quality Report"
        ws_meta['A1'].font = title_font
        ws_meta['A2'] = f"Generated: {timestamp}"
        ws_meta['A2'].font = subtitle_font

        # Metadata table
        row = 4
        headers = ["Field", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws_meta.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Metadata values
        meta_data = [
            ("Run ID", run_id),
            ("Timestamp", timestamp),
            ("Artifact", artifact),
            ("Criteria", criteria),
            ("Iterations", f"{iterations} / {max_iterations}"),
            ("Priority Profile", priority_profile),
            ("Overall Score", f"{overall_score:.3f}"),
            ("Status", "PASSED" if passed else "FAILED"),
            ("Gate Status", "PASSED" if gate_status is True else "FAILED" if gate_status is False else "NOT_EVALUATED"),
            ("Gate Policy", gate_policy),
            ("Benchmark Percentile", f"{benchmark_percentile:.1f}%" if benchmark_percentile else "N/A"),
            ("Benchmark Comparison", benchmark_comparison or "N/A"),
        ]

        for idx, (field, value) in enumerate(meta_data, 1):
            # Convert non-scalar values to strings for Excel compatibility
            if isinstance(value, (list, tuple)):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            ws_meta.cell(row=row+idx, column=1, value=field).border = border
            ws_meta.cell(row=row+idx, column=2, value=value).border = border

        # Auto-adjust column widths
        ws_meta.column_dimensions['A'].width = 20
        ws_meta.column_dimensions['B'].width = 35

        # ============================================================================
        # SHEET 2: Summary
        # ============================================================================
        ws_summary = wb.create_sheet("Summary")

        # Title
        ws_summary['A1'] = "Quality Score Summary"
        ws_summary['A1'].font = title_font

        # Summary data
        summary_data = [
            ["Metric", "Value"],
            ["Overall Score", f"{overall_score:.3f}"],
            ["Status", "PASSED" if passed else "FAILED"],
            ["Total Categories", len(categories)],
            ["Iterations Completed", iterations],
            ["Max Iterations", max_iterations],
        ]

        # Count total rules
        total_passed = sum(c.get("passed", 0) for c in categories.values())
        total_failed = sum(c.get("failed", 0) for c in categories.values())
        total_warnings = sum(c.get("warnings", 0) for c in categories.values())
        total_rules = total_passed + total_failed

        summary_data.extend([
            ["Total Rules", total_rules],
            ["Passed Rules", total_passed],
            ["Failed Rules", total_failed],
            ["Warnings", total_warnings],
        ])

        # Write summary table
        row = 3
        for r_idx, row_data in enumerate(summary_data):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row+r_idx, column=c_idx, value=value)
                if r_idx == 0:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                else:
                    cell.border = border
                    if c_idx == 1:  # Metric column
                        cell.font = Font(bold=True)

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20

        # ============================================================================
        # SHEET 3: Category Scores with Conditional Formatting
        # ============================================================================
        ws_scores = wb.create_sheet("Category Scores")

        # Title
        ws_scores['A1'] = "Category Scores Detail"
        ws_scores['A1'].font = title_font

        # Headers
        headers = ["Category", "Score", "Weighted Score", "Passed", "Failed", "Warnings", "Total", "Priority", "Weight"]
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws_scores.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write OVERALL row first
        ws_scores.cell(row=row+1, column=1, value="OVERALL").font = Font(bold=True)
        ws_scores.cell(row=row+1, column=2, value=overall_score).font = Font(bold=True, color="008000" if overall_score >= 0.8 else "FF0000")
        ws_scores.cell(row=row+1, column=3, value=overall_score).font = Font(bold=True)
        for col in range(1, 10):
            ws_scores.cell(row=row+1, column=col).border = border

        # Category rows
        score_row = row + 2
        scores_for_formatting = []

        for category_name, category_data in categories.items():
            score = category_data.get("score", 0.0)
            passed_rules = category_data.get("passed", 0)
            total_rules = category_data.get("total", 0)
            failed_rules = category_data.get("failed", 0)
            warnings = category_data.get("warnings", 0)
            priority = category_data.get("priority", "medium")
            weight = category_data.get("weight", 1.0)

            ws_scores.cell(row=score_row, column=1, value=category_name).border = border
            ws_scores.cell(row=score_row, column=2, value=round(score, 3)).border = border
            ws_scores.cell(row=score_row, column=3, value=round(score * weight, 3)).border = border
            ws_scores.cell(row=score_row, column=4, value=passed_rules).border = border
            ws_scores.cell(row=score_row, column=5, value=failed_rules).border = border
            ws_scores.cell(row=score_row, column=6, value=warnings).border = border
            ws_scores.cell(row=score_row, column=7, value=total_rules).border = border
            ws_scores.cell(row=score_row, column=8, value=priority).border = border
            ws_scores.cell(row=score_row, column=9, value=weight).border = border

            # Collect score cells for conditional formatting
            scores_for_formatting.append(ws_scores.cell(row=score_row, column=2))

            score_row += 1

        # Auto-adjust column widths
        ws_scores.column_dimensions['A'].width = 20
        ws_scores.column_dimensions['B'].width = 12
        ws_scores.column_dimensions['C'].width = 15
        ws_scores.column_dimensions['D'].width = 10
        ws_scores.column_dimensions['E'].width = 10
        ws_scores.column_dimensions['F'].width = 10
        ws_scores.column_dimensions['G'].width = 10
        ws_scores.column_dimensions['H'].width = 12
        ws_scores.column_dimensions['I'].width = 10

        # Conditional formatting for scores (color scale: red -> yellow -> green)
        if scores_for_formatting:
            color_scale = ColorScaleRule(
                start_type='min', start_color='FF0000',
                mid_type='percentile', mid_value=50, mid_color='FFFF00',
                end_type='max', end_color='008000'
            )
            ws_scores.conditional_formatting.add(f'B{row+2}:B{score_row}', color_scale)

        # Freeze header row
        ws_scores.freeze_panes = f'A{row+1}'

        # ============================================================================
        # SHEET 4: Failed Rules Details
        # ============================================================================
        ws_rules = wb.create_sheet("Failed Rules")

        # Title
        ws_rules['A1'] = "Failed Rules Details"
        ws_rules['A1'].font = title_font

        # Headers
        headers = ["Run ID", "Rule ID", "Category", "Severity", "Message", "Suggestion"]
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws_rules.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Failed rules data
        failed_rules = evaluation.get("failed_rules", [])
        rule_row = row + 1

        for rule in failed_rules:
            severity = rule.get("severity", "fail")
            ws_rules.cell(row=rule_row, column=1, value=run_id).border = border
            ws_rules.cell(row=rule_row, column=2, value=rule.get("id", rule.get("rule_id", "unknown"))).border = border
            ws_rules.cell(row=rule_row, column=3, value=rule.get("category", "general")).border = border

            severity_cell = ws_rules.cell(row=rule_row, column=4, value=severity.upper())
            severity_cell.border = border
            if severity == "fail":
                severity_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif severity == "warning":
                severity_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            ws_rules.cell(row=rule_row, column=5, value=rule.get("message", "")).border = border
            ws_rules.cell(row=rule_row, column=6, value=rule.get("suggestion", "")).border = border

            rule_row += 1

        # Auto-adjust column widths
        ws_rules.column_dimensions['A'].width = 12
        ws_rules.column_dimensions['B'].width = 20
        ws_rules.column_dimensions['C'].width = 15
        ws_rules.column_dimensions['D'].width = 12
        ws_rules.column_dimensions['E'].width = 50
        ws_rules.column_dimensions['F'].width = 50

        # Freeze header row
        ws_rules.freeze_panes = f'A{row+1}'

        # ============================================================================
        # Save to bytes
        # ============================================================================
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def _generate_excel_safe(self, result: Dict[str, Any]) -> Optional[bytes]:
        """Generate Excel report with fallback to CSV if openpyxl not available

        Args:
            result: Quality loop result dict

        Returns:
            Excel bytes or None if openpyxl not available
        """
        try:
            return self._generate_excel(result)
        except ImportError:
            # openpyxl not available, return None
            return None


def export_quality_reports(
    result: Dict[str, Any],
    previous_score: Optional[float] = None,
    formats: Optional[List[ReportFormat]] = None,
    output_dir: Optional[str] = None,
    filename_prefix: str = "quality_report",
    include_timeline: bool = True,
    include_details: bool = True,
    compact_console: bool = False,
    console_theme: str = "default",
    json_pretty: bool = True,
) -> ExportResult:
    """Convenience function to export quality reports in multiple formats

    Args:
        result: Quality loop result dict from QualityLoop.run()
        previous_score: Previous run score for trend calculation
        formats: List of formats to generate (default: all)
                 Options: "console", "json", "html", "markdown", "csv", "excel" (Exp 106)
        output_dir: Directory to save reports (None = in-memory only)
        filename_prefix: Prefix for generated filenames
        include_timeline: Include timeline charts in HTML/Markdown
        include_details: Include detailed rule breakdowns
        compact_console: Use compact format for console output
        console_theme: Color theme for console (default, dark, high-contrast, minimal)
        json_pretty: Pretty-print JSON output

    Returns:
        ExportResult with all generated reports

    Example:
        >>> result = loop.run(artifact="spec.md", criteria_name="backend")
        >>> export_result = export_quality_reports(
        ...     result,
        ...     formats=["json", "html", "excel"],  # Excel with conditional formatting
        ...     output_dir="./reports"
        ... )
        >>> print(export_result.get_console_output())
        >>> html_path = export_result.reports["html"].path

    Note:
        Excel export requires openpyxl: pip install openpyxl
        Excel files include multiple sheets: Metadata, Summary, Category Scores, Failed Rules
    """
    config = ExportConfig(
        formats=formats,
        output_dir=output_dir,
        filename_prefix=filename_prefix,
        include_timeline=include_timeline,
        include_details=include_details,
        compact_console=compact_console,
        console_theme=console_theme,
        json_pretty=json_pretty,
    )

    exporter = ReportExporter(config)
    return exporter.export(result, previous_score)


def export_result_card_json(
    result: Dict[str, Any],
    output_path: Optional[str] = None,
    pretty: bool = True,
) -> str:
    """Export result card data as JSON for programmatic access

    Args:
        result: Quality loop result dict
        output_path: Optional file path to save JSON
        pretty: Pretty-print JSON with indentation

    Returns:
        JSON string with result card data

    Example:
        >>> result = loop.run(artifact="spec.md", criteria_name="backend")
        >>> card_json = export_result_card_json(result, "result_card.json")
        >>> data = json.loads(card_json)
        >>> print(f"Status: {data['status']}, Score: {data['score']}")
    """
    card_data = create_result_card_data(result)

    card_dict = {
        "status": card_data.status.value,
        "score": card_data.score,
        "passed": card_data.passed,
        "iteration": {
            "current": card_data.iteration,
            "max": card_data.max_iterations,
        },
        "phase": card_data.phase,
        "rules": {
            "total": card_data.total_rules,
            "passed": card_data.passed_rules,
            "failed": card_data.failed_rules,
            "warnings": card_data.warnings,
        },
        "duration_seconds": card_data.duration_seconds,
        "priority_profile": card_data.priority_profile,
        "trend_change": card_data.trend_change,
        "gate_status": card_data.gate_status,
        "categories": [
            {
                "category": cat.category,
                "failed": cat.failed_count,
                "warnings": cat.warning_count,
                "total": cat.total_count,
                "priority": cat.priority,
                "sample_rules": cat.sample_rules,
            }
            for cat in card_data.category_summaries
        ],
        "action_items": [
            {
                "priority": item.priority,
                "title": item.title,
                "command": item.command,
                "description": item.description,
            }
            for item in card_data.action_items
        ],
        "generated_at": datetime.now().isoformat(),
    }

    json_content = json.dumps(card_dict, indent=2 if pretty else None, ensure_ascii=False)

    if output_path:
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        output_file.write_text(json_content, encoding="utf-8")

    return json_content


def format_export_summary(export_result: ExportResult) -> str:
    """Format a summary of exported reports

    Args:
        export_result: Export result from ReportExporter.export()

    Returns:
        Formatted summary string
    """
    lines = [
        "# Quality Report Export Summary",
        "",
        f"**Generated:** {export_result.generated_at}",
        f"**Total Size:** {export_result.total_size_bytes:,} bytes",
        "",
        "## Exported Formats",
        "",
    ]

    if not export_result.reports:
        lines.append("No reports generated.")
    else:
        for format, report in export_result.reports.items():
            path_str = f" → `{report.path}`" if report.path else " (in-memory)"
            lines.append(f"- **{format.upper()}**: {report.size_bytes:,} bytes{path_str}")

    return "\n".join(lines)


# ============================================================
# Markdown Report Generator (formerly markdown_report.py)
# ============================================================

class MarkdownReportGenerator:
    """Generate Markdown reports for quality evaluation results"""

    def generate(
        self,
        result: Dict[str, Any],
        output_path: Optional[str] = None,
        include_timeline: bool = True,
        include_details: bool = True,
    ) -> str:
        """Generate Markdown report from evaluation result"""
        state = result.get("state", {})
        score = result.get("score", 0.0)
        passed = result.get("passed", False)
        priority_profile = result.get("priority_profile", "default")
        gate_result = result.get("gate_result")

        md_parts = [
            self._get_header(state, score, passed, priority_profile),
            self._get_summary_section(state, score, passed),
        ]

        if gate_result:
            md_parts.append(self._get_gate_section(gate_result))

        if include_timeline:
            md_parts.append(self._get_timeline_section(state))

        if include_details:
            md_parts.append(self._get_details_section(state))

        md_parts.append(self._get_footer())
        md_content = "\n\n".join(md_parts)

        if output_path:
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            output_file.write_text(md_content, encoding="utf-8")

        return md_content

    def _get_header(self, state: Dict, score: float, passed: bool, priority_profile: str) -> str:
        run_id = state.get("run_id", "unknown")
        task_alias = state.get("task_alias", "unknown")
        status_text = "PASSED" if passed else "FAILED"

        return f"""# Spec Kit Quality Report

**Task:** `{task_alias}` | **Run:** `{run_id}` | **Profile:** `{priority_profile}`

## Score: `{score:.2f}` / 1.0

**Status:** **{status_text}**
"""

    def _get_summary_section(self, state: Dict, score: float, passed: bool) -> str:
        iteration = state.get("iteration", 1)
        max_iterations = state.get("max_iterations", 4)
        phase = state.get("phase", "A")
        started_at = state.get("started_at", "")
        progress = (iteration / max_iterations) * 100
        category_breakdown = self._get_category_breakdown(state)

        summary = f"""## Summary

| Metric | Value |
|--------|-------|
| **Iteration** | {iteration} / {max_iterations} ({progress:.0f}%) |
| **Phase** | {phase} |
| **Quality Score** | {score:.2f} |
| **Status** | {'PASS' if passed else 'FAIL'} |

{"**Started:** " + started_at if started_at else ""}
"""

        if category_breakdown:
            summary += "\n\n" + category_breakdown

        return summary

    def _get_gate_section(self, gate_result: Dict[str, Any]) -> str:
        gate_status = gate_result.get("gate_result", "unknown").upper()
        passed = gate_result.get("passed", False)
        policy_name = gate_result.get("policy_name", "unknown")
        policy_desc = gate_result.get("policy_description", "")
        overall_threshold = gate_result.get("overall_threshold", 0.0)
        overall_score = gate_result.get("overall_score", 0.0)
        messages = gate_result.get("messages", [])
        severity_counts = gate_result.get("severity_counts", {})
        category_scores = gate_result.get("category_scores", {})
        block_on_failure = gate_result.get("block_on_failure", True)

        severity_lines = []
        for sev in ["critical", "high", "medium", "low", "info"]:
            count = severity_counts.get(sev, 0)
            if count > 0:
                severity_lines.append(f"| {sev.capitalize()} | {count} |")

        category_lines = []
        for cat, cat_score in sorted(category_scores.items(), key=lambda x: x[1], reverse=True):
            percentage = cat_score * 100
            category_lines.append(f"| {cat} | {percentage:.0f}% |")

        messages_md = ""
        if messages:
            messages_md = "\n### Gate Violations\n\n"
            for i, msg in enumerate(messages[:10], 1):
                messages_md += f"{i}. {msg}\n"
            if len(messages) > 10:
                messages_md += f"\n... and {len(messages) - 10} more violations\n"

        gate_section = f"""## Quality Gate Results

| | |
|---|---|
| **Policy** | {policy_name} |
| **Description** | {policy_desc} |
| **Status** | **{gate_status}** |
| **Score** | {overall_score:.2f} / {overall_threshold:.2f} |
| **Blocking** | {'Yes' if block_on_failure else 'No'} |
"""

        if severity_lines:
            gate_section += "\n### Issues by Severity\n\n"
            gate_section += "| Severity | Count |\n|----------|-------|\n"
            gate_section += "\n".join(severity_lines) + "\n"

        if category_lines:
            gate_section += "\n### Category Scores\n\n"
            gate_section += "| Category | Score |\n|----------|-------|\n"
            gate_section += "\n".join(category_lines) + "\n"

        if messages_md:
            gate_section += messages_md
        elif passed:
            gate_section += "\nAll gate checks passed!\n"

        return gate_section

    def _get_category_breakdown(self, state: Dict) -> str:
        evaluation = state.get("evaluation", {})
        failed_rules = evaluation.get("failed_rules", [])
        warnings = evaluation.get("warnings", [])

        from collections import defaultdict
        category_stats = defaultdict(lambda: {"fail": 0, "warn": 0})

        for rule in failed_rules:
            category = rule.get("category", "general")
            category_stats[category]["fail"] += 1

        for rule in warnings:
            category = rule.get("category", "general")
            category_stats[category]["warn"] += 1

        if not category_stats:
            return ""

        rows = []
        total_issues = sum(cat["fail"] + cat["warn"] for cat in category_stats.values())

        for category, stats in sorted(category_stats.items(), key=lambda x: x[1]["fail"] + x[1]["warn"], reverse=True):
            fail_count = stats["fail"]
            warn_count = stats["warn"]
            total = fail_count + warn_count
            percentage = (total / total_issues * 100) if total_issues > 0 else 0
            rows.append(f"| **{category}** | {fail_count} | {warn_count} | {total} | {percentage:.1f}% |")

        header = "| Category | Failed | Warnings | Total | % |\n|----------|--------|----------|-------|-------|"

        return f"""## Issues by Category

{header}
{chr(10).join(rows)}
"""

    def _get_timeline_section(self, state: Dict) -> str:
        events = self._extract_score_events(state)

        rows = []
        for event in events:
            label = event["label"]
            score_val = event["score"]
            percentage = score_val * 100
            rows.append(f"| {label} | {percentage:.1f}% |")

        header = "| Iteration | Score |\n|-----------|-------|"

        return f"""## Score Timeline

{header}
{chr(10).join(rows)}
"""

    def _get_details_section(self, state: Dict) -> str:
        evaluation = state.get("evaluation", {})
        failed_rules = evaluation.get("failed_rules", [])
        warnings = evaluation.get("warnings", [])

        from collections import defaultdict
        category_rules = defaultdict(lambda: {"fail": [], "warn": []})

        for rule in failed_rules:
            category = rule.get("category", "general")
            category_rules[category]["fail"].append(rule)

        for rule in warnings:
            category = rule.get("category", "general")
            category_rules[category]["warn"].append(rule)

        if not category_rules:
            return """## Details

### Failed Rules & Warnings

*No issues found!*
"""

        sections = []
        for category in sorted(category_rules.keys(), key=lambda c: sum(len(r) for r in category_rules[c].values()), reverse=True):
            fail_rules = category_rules[category]["fail"]
            warn_rules = category_rules[category]["warn"]

            if not fail_rules and not warn_rules:
                continue

            sections.append(f"\n### {category.title()}\n")

            if fail_rules:
                sections.append("\n#### Failed Rules\n\n")
                for i, rule in enumerate(fail_rules, 1):
                    rule_id = rule.get("rule_id", "unknown")
                    reason = rule.get("reason", "No reason provided")
                    sections.append(f"**{i}. `{rule_id}`**\n\n{reason}\n")

            if warn_rules:
                sections.append("\n#### Warnings\n\n")
                for i, rule in enumerate(warn_rules, 1):
                    rule_id = rule.get("rule_id", "unknown")
                    reason = rule.get("reason", "No reason provided")
                    sections.append(f"**{i}. `{rule_id}`**\n\n{reason}\n")

        return f"""## Details

### Failed Rules & Warnings by Category

{''.join(sections)}
"""

    def _get_footer(self) -> str:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return f"""---

*Generated by Spec Kit Quality Loop at {timestamp}*
"""

    def _extract_score_events(self, state: Dict) -> List[Dict[str, Any]]:
        events = [{"label": "Start", "score": 0.0}]

        current_score = state.get("current_score", 0.0)
        iteration = state.get("iteration", 1)
        events.append({"label": f"Iteration {iteration}", "score": current_score})

        last_score = state.get("last_score")
        if last_score is not None and abs(last_score - current_score) > 0.01:
            events.insert(-1, {"label": f"Iteration {iteration - 1}", "score": last_score})

        return events


def generate_markdown_report(
    result: Dict[str, Any],
    output_path: Optional[str] = None,
    include_timeline: bool = True,
    include_details: bool = True,
) -> str:
    """Convenience function to generate Markdown report"""
    generator = MarkdownReportGenerator()
    return generator.generate(result, output_path, include_timeline, include_details)
